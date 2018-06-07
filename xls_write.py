#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time:2018/5/2
# @File:xls_write.py

import os
import time
import split_txt as st
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Font, PatternFill, Alignment

# 读取excle模板
xls_path = os.path.join(os.getcwd(), '日报模板.xlsx')
xls_wb = load_workbook(xls_path)
xls_ws = xls_wb.worksheets[0]

#将列表的内容依次填入表格中
row_pick = 2
for cell_val in st.sorted_content:
    cell = xls_ws.cell(row=row_pick,column=4)
    cell.value = cell_val
    row_pick += 1

# 填入结果后，有些单元格格式（线框，字体，填充颜色）有变化，需进行适当调整
cell_bd = NamedStyle(name="cell_bd")
bd = Side(style='thin', color='000000')
cell_bd.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# 设置单元格左对齐居中，自动换行
al_all = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_col5 = Alignment(horizontal="center", vertical="center", wrap_text=True)
al_col4 = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 调整单元格线框，自动换行（2-119行，1-7列区间）
for row in range(2, 120):
    for col in range(1, 8):
        xls_ws.cell(row, col).style = cell_bd
        xls_ws.cell(row, col).alignment = al_all

# 调整第五列为居中对齐
for row in range(2, 120):
    xls_ws.cell(row=row, column=5).alignment = al_col5

# 调整第四列为顶端左对齐
for row in range(2, 119):
    xls_ws.cell(row=row, column=4).alignment = al_col4

# 调整单元格字体
for col in range(1, 8):
    xls_ws.cell(row=1, column=col).font = Font(size=10, bold=True)

for row in range(1, 120):
    xls_ws.cell(row=row, column=1).font = Font(size=10, bold=True)

for row in range(2, 120):
    xls_ws.cell(row=row, column=2).font = Font(size=12)

for row in range(2, 120):
    for col in range(3, 8):
        xls_ws.cell(row=row, column=col).font = Font(size=9)
        if col == 4:
            xls_ws.cell(row=row,column=col).font = Font(color="0000FF",size=9)

# 单元格C118,F101,G101颜色填充设置
cell_fl = PatternFill(fgColor='FFFF00', patternType='solid', fill_type='solid')
xls_ws.cell(row=118, column=3).fill = cell_fl
xls_ws.cell(row=101, column=6).fill = cell_fl
xls_ws.cell(row=101, column=7).fill = cell_fl

# 网状网结果填写
al_wzw = Alignment(horizontal="left", vertical="bottom")
xls_ws['D119'] = st.wzw_status
xls_ws.cell(row=119, column=4).alignment = al_wzw

# 调整单元格格式为百分比
row = 2
while True:
    if row > 97:
        break
    for count in range(3):
        row += 1
        cell_percentage = xls_ws.cell(row=row, column=4)
        cell_percentage.number_format = '0.000%'
    row += 1

# 文件名格式化
hour_now = time.strftime('%H', time.localtime(time.time()))
if int(hour_now) > 12:
    am_pm = "下午"
else:
    am_pm = "上午"
file_name = ("物联网内容计费主机巡检日报" + "_" + time.strftime("%Y%m%d", time.localtime(time.time()))
             + am_pm + ".xlsx")

xls_wb.save(file_name)
